#!/usr/bin/env python3
"""
RYNO SEO Tools – Keyword Report Generator (Web App)
====================================================
Run:  python ryno_report_app.py
Then open: http://localhost:5050

Requirements:
    pip install flask pandas openpyxl
"""

from flask import Flask, request, send_file, render_template_string
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
import io, os, tempfile

app = Flask(__name__)

# ─── Color palette ─────────────────────────────────────────────────────────
DARK_BG     = "1A2332"
ACCENT_BLUE = "2563EB"
ACCENT_TEAL = "0EA5E9"
LIGHT_BLUE  = "EFF6FF"
MID_BLUE    = "BFDBFE"
WHITE       = "FFFFFF"
DARK_TEXT   = "1E293B"
GRAY_BG     = "F8FAFC"
GRAY_BORDER = "CBD5E1"
GREEN_UP    = "16A34A"
RED_DOWN    = "DC2626"
AMBER       = "D97706"
GREEN_LIGHT = "DCFCE7"
RED_LIGHT   = "FEE2E2"
AMBER_LIGHT = "FEF3C7"
PURPLE      = "7C3AED"
SLATE       = "94A3B8"

def side(color=GRAY_BORDER, style="thin"):
    return Side(border_style=style, color=color)

def full_border(color=GRAY_BORDER):
    s = side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def hfont(size=11, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def bfont(size=10, bold=False, color=DARK_TEXT):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def rank_style(val):
    if val == 1:    return ("16A34A", WHITE, True)
    elif val <= 3:  return ("22C55E", WHITE, True)
    elif val <= 10: return (GREEN_LIGHT, "15803D", False)
    elif val <= 20: return (AMBER_LIGHT, "92400E", False)
    elif val <= 50: return ("FEF9C3", "713F12", False)
    else:           return (RED_LIGHT, "991B1B", False)


def client_from_filename(filename):
    """Extract client slug from filenames like ryno_acehardwarehome__123__....csv"""
    import re
    base = os.path.splitext(os.path.basename(str(filename)))[0]
    parts = base.split('__')
    slug = re.sub(r'^ryno_', '', parts[0], flags=re.IGNORECASE)
    return slug if slug else base


def generate_excel(df, original_filename=''):
    df['date'] = pd.to_datetime(df['requested']).dt.date
    dates     = sorted(df['date'].unique())
    locations = list(df['location'].unique())

    # Prefer result__url__domain column; fall back to client slug from filename
    client_domain = ""
    if 'result__url__domain' in df.columns:
        domains = df['result__url__domain'].dropna().unique()
        if len(domains):
            client_domain = str(domains[0]).strip()
    if not client_domain and original_filename:
        client_domain = client_from_filename(original_filename)

    wb = Workbook()
    latest_date = dates[-1]
    prior_date  = dates[-2] if len(dates) > 1 else dates[-1]
    latest_df   = df[df['date'] == latest_date]
    prior_df    = df[df['date'] == prior_date]
    avg_rank_latest = latest_df['top_rank__avg__total__value'].mean()
    top10_count = (latest_df['top_rank__avg__total__value'] <= 10).sum()

    # ── DASHBOARD ──────────────────────────────────────────────────────────
    dash = wb.active
    dash.title = "📊 Dashboard"
    dash.sheet_properties.tabColor = ACCENT_BLUE

    dash.row_dimensions[1].height = 60
    for col in range(1, 22):
        dash.cell(row=1, column=col).fill = fill(DARK_BG)
    dash.merge_cells("A1:U1")
    h1 = dash["A1"]
    h1.value = "🔍  RYNO SEO Tools – Keyword Ranking Report"
    h1.font = Font(name="Arial", size=20, bold=True, color=WHITE)
    h1.alignment = center()

    dash.row_dimensions[2].height = 28
    dash.merge_cells("A2:U2")
    h2 = dash["A2"]
    h2.value = (f"Client: {client_domain}   |   "
                f"{dates[0].strftime('%b %d, %Y')} → {dates[-1].strftime('%b %d, %Y')}"
                f"   |   {len(locations)} Locations   |   {len(dates)} Weeks")
    h2.font = Font(name="Arial", size=11, color=MID_BLUE)
    h2.fill = fill(DARK_BG)
    h2.alignment = center()

    dash.row_dimensions[3].height = 14

    kpi_defs = [
        ("🔑", "Total Keywords",     str(df['phrase'].nunique() * len(locations)), ACCENT_BLUE),
        ("📍", "Locations Tracked",  str(len(locations)),                          "0891B2"),
        ("📅", "Tracking Weeks",     str(len(dates)),                              PURPLE),
        ("🏆", "Avg Rank (Latest)",  f"{avg_rank_latest:.1f}",                    AMBER if avg_rank_latest > 20 else GREEN_UP),
        ("⭐", "Keywords Rank 1–10", str(top10_count),                             GREEN_UP),
    ]
    col_starts = [1, 4, 7, 10, 13]
    for i, (icon, label, val, color) in enumerate(kpi_defs):
        cs, ce = col_starts[i], col_starts[i] + 2
        for r in range(4, 8):
            dash.row_dimensions[r].height = [18, 50, 30, 16][r - 4]
            dash.merge_cells(start_row=r, start_column=cs, end_row=r, end_column=ce)
            for c in range(cs, ce + 1):
                dash.cell(row=r, column=c).fill = fill(color)
        dash.cell(row=4, column=cs).value = icon
        dash.cell(row=4, column=cs).font = Font(size=18)
        dash.cell(row=4, column=cs).alignment = center()
        dash.cell(row=5, column=cs).value = val
        dash.cell(row=5, column=cs).font = Font(name="Arial", size=26, bold=True, color=WHITE)
        dash.cell(row=5, column=cs).alignment = center()
        dash.cell(row=6, column=cs).value = label
        dash.cell(row=6, column=cs).font = Font(name="Arial", size=10, color="E2E8F0")
        dash.cell(row=6, column=cs).alignment = center()

    dash.row_dimensions[9].height = 28
    dash.merge_cells("A9:U9")
    sec = dash["A9"]
    sec.value = "  📍  Average Rank by Location – Latest vs. Prior Week"
    sec.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    sec.fill = fill(ACCENT_BLUE)
    sec.alignment = left_align()

    tbl_headers = ["Location", "Latest Rank", "Prior Rank", "Change", "Trend", "# Top 10", "# Top 20"]
    tbl_widths  = [32, 14, 14, 10, 14, 11, 11]
    for c_idx, (h_txt, w) in enumerate(zip(tbl_headers, tbl_widths), start=1):
        dash.column_dimensions[get_column_letter(c_idx)].width = w
        cell = dash.cell(row=10, column=c_idx)
        cell.value = h_txt
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill = fill(DARK_BG)
        cell.alignment = center()
        cell.border = full_border(DARK_BG)

    for row_i, loc in enumerate(locations, start=11):
        lat = latest_df[latest_df['location'] == loc]['top_rank__avg__total__value']
        pri = prior_df[prior_df['location'] == loc]['top_rank__avg__total__value']
        avg_lat = lat.mean() if len(lat) else 0
        avg_pri = pri.mean() if len(pri) else 0
        change  = avg_lat - avg_pri
        t10 = (lat <= 10).sum()
        t20 = (lat <= 20).sum()
        trend       = "▲ Improved" if change < 0 else ("▼ Dropped" if change > 0 else "→ Stable")
        trend_color = GREEN_UP if change < 0 else (RED_DOWN if change > 0 else AMBER)
        row_bg = WHITE if row_i % 2 == 1 else GRAY_BG
        vals = [loc, f"{avg_lat:.1f}", f"{avg_pri:.1f}", f"{change:+.1f}", trend, t10, t20]
        for c_idx, val in enumerate(vals, start=1):
            cell = dash.cell(row=row_i, column=c_idx)
            cell.value = val
            cell.font = bfont(color=(trend_color if c_idx == 5 else DARK_TEXT), bold=(c_idx == 5))
            cell.fill = fill(row_bg)
            cell.alignment = center() if c_idx != 1 else left_align()
            cell.border = full_border()
        dash.row_dimensions[row_i].height = 20

    note_row = 11 + len(locations) + 1
    dash.row_dimensions[note_row].height = 44
    dash.merge_cells(f"A{note_row}:U{note_row}")
    note = dash.cell(row=note_row, column=1)
    note.value = ("💡  HOW TO READ:  Lower rank = better (Rank 1 = #1 on Google).  "
                  "▲ Improved = rank went DOWN (moved up).  Rank 101 = not in top 100.  "
                  "Each tab = one city with full keyword history, ranking URLs, and charts.")
    note.font = Font(name="Arial", size=10, italic=True, color="64748B")
    note.fill = fill(LIGHT_BLUE)
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note.border = Border(left=side(ACCENT_BLUE, "medium"), top=side(), bottom=side(), right=side())

    for col in range(8, 22):
        dash.column_dimensions[get_column_letter(col)].width = 8

    # ── LOCATION SHEETS ────────────────────────────────────────────────────
    has_url = 'result__url__url' in df.columns

    for loc in locations:
        city     = loc.split(",")[0]
        tab_name = f"📍 {city}"[:31]
        ws = wb.create_sheet(title=tab_name)
        ws.sheet_properties.tabColor = ACCENT_TEAL

        loc_df   = df[df['location'] == loc].copy()
        keywords = sorted(loc_df['phrase'].unique())

        pivot_rank = loc_df.pivot_table(
            index='phrase', columns='date',
            values='top_rank__avg__total__value', aggfunc='first'
        ).reindex(columns=dates)

        url_map = {}
        if has_url:
            # Scan all dates newest-first; grab the most recent non-null URL per keyword
            for check_date in reversed(dates):
                date_rows = loc_df[loc_df['date'] == check_date]
                for _, row in date_rows.iterrows():
                    kw  = row['phrase']
                    url = row.get('result__url__url', '')
                    if kw not in url_map and pd.notna(url) and str(url).startswith('http'):
                        url_map[kw] = str(url)

        total_cols = 1 + len(dates) + (1 if has_url else 0) + 3
        ws.row_dimensions[1].height = 50
        ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
        c = ws["A1"]
        c.value = f"📍  {loc}  –  Keyword Ranking History"
        c.font = Font(name="Arial", size=16, bold=True, color=WHITE)
        c.fill = fill(DARK_BG)
        c.alignment = center()

        ws.row_dimensions[2].height = 24
        ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
        c2 = ws["A2"]
        c2.value = (f"{len(keywords)} keywords  |  {len(dates)} weeks  |  "
                    f"Lower rank = better  |  Rank 101 = not in top 100")
        c2.font = Font(name="Arial", size=10, color=SLATE)
        c2.fill = fill(DARK_BG)
        c2.alignment = center()

        ws.row_dimensions[3].height = 12
        ws.row_dimensions[4].height = 48
        ws.column_dimensions["A"].width = 38

        c = ws.cell(row=4, column=1)
        c.value = "Keyword"
        c.font = hfont()
        c.fill = fill(ACCENT_BLUE)
        c.alignment = center()
        c.border = full_border(ACCENT_BLUE)

        for col_i, d in enumerate(dates, start=2):
            ws.column_dimensions[get_column_letter(col_i)].width = 13
            c = ws.cell(row=4, column=col_i)
            c.value = d.strftime("%b %d\n%Y")
            c.font = hfont(size=9)
            c.fill = fill(ACCENT_BLUE)
            c.alignment = center()
            c.border = full_border(ACCENT_BLUE)

        url_col   = len(dates) + 2 if has_url else None
        chg_col   = (url_col + 1) if url_col else (len(dates) + 2)
        trend_col = chg_col + 1
        best_col  = trend_col + 1

        if has_url:
            ws.column_dimensions[get_column_letter(url_col)].width = 52
            c = ws.cell(row=4, column=url_col)
            c.value = "Ranking URL (Latest)"
            c.font = hfont(size=9)
            c.fill = fill(ACCENT_TEAL)
            c.alignment = center()
            c.border = full_border(ACCENT_TEAL)

        for col_i, (label, clr) in [(chg_col, ("Δ Change", PURPLE)),
                                     (trend_col, ("Trend", PURPLE)),
                                     (best_col, ("Best Rank", PURPLE))]:
            ws.column_dimensions[get_column_letter(col_i)].width = 13 if col_i != trend_col else 15
            c = ws.cell(row=4, column=col_i)
            c.value = label
            c.font = hfont(size=9)
            c.fill = fill(clr)
            c.alignment = center()
            c.border = full_border(clr)

        for row_i, kw in enumerate(keywords, start=5):
            ws.row_dimensions[row_i].height = 20
            row_bg = WHITE if row_i % 2 == 0 else GRAY_BG

            c = ws.cell(row=row_i, column=1)
            c.value = kw
            c.font = bfont()
            c.fill = fill(row_bg)
            c.alignment = left_align()
            c.border = full_border()

            ranks = []
            for col_i, d in enumerate(dates, start=2):
                val = pivot_rank.loc[kw, d] if kw in pivot_rank.index and d in pivot_rank.columns else None
                c = ws.cell(row=row_i, column=col_i)
                if pd.notna(val):
                    int_val = int(val)
                    ranks.append(int_val)
                    bg, fg, bold = rank_style(int_val)
                    c.value = int_val
                    c.fill = fill(bg)
                    c.font = Font(name="Arial", size=10, bold=bold, color=fg)
                else:
                    c.value = "–"
                    c.fill = fill(row_bg)
                    c.font = bfont(color=SLATE)
                c.alignment = center()
                c.border = full_border()

            if has_url:
                url_val = url_map.get(kw, "")
                c = ws.cell(row=row_i, column=url_col)
                c.value = url_val if url_val else "—"
                if url_val:
                    c.hyperlink = url_val
                    c.font = Font(name="Arial", size=9, color="1D4ED8", underline="single")
                else:
                    c.font = bfont(color=SLATE, size=9)
                c.fill = fill(row_bg)
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                c.border = full_border()

            first_r = ranks[0] if ranks else None
            last_r  = ranks[-1] if ranks else None

            c = ws.cell(row=row_i, column=chg_col)
            if first_r and last_r:
                change = last_r - first_r
                c.value = change
                if change < 0:
                    c.font = Font(name="Arial", size=10, bold=True, color=GREEN_UP)
                    c.fill = fill(GREEN_LIGHT)
                elif change > 0:
                    c.font = Font(name="Arial", size=10, bold=True, color=RED_DOWN)
                    c.fill = fill(RED_LIGHT)
                else:
                    c.font = bfont()
                    c.fill = fill(row_bg)
            else:
                c.value = "–"
                c.font = bfont(color=SLATE)
                c.fill = fill(row_bg)
            c.alignment = center()
            c.border = full_border()

            c = ws.cell(row=row_i, column=trend_col)
            if first_r and last_r:
                change = last_r - first_r
                if change < -10:   t, tb, tf = "🚀 Big Gain",  GREEN_LIGHT, GREEN_UP
                elif change < 0:   t, tb, tf = "▲ Improving",  GREEN_LIGHT, GREEN_UP
                elif change == 0:  t, tb, tf = "→ Stable",     AMBER_LIGHT, AMBER
                elif change <= 10: t, tb, tf = "▼ Slipping",   RED_LIGHT,   RED_DOWN
                else:              t, tb, tf = "⚠ Big Drop",   RED_LIGHT,   RED_DOWN
                c.value = t
                c.font = Font(name="Arial", size=10, bold=True, color=tf)
                c.fill = fill(tb)
            else:
                c.value = "–"
                c.font = bfont(color=SLATE)
                c.fill = fill(row_bg)
            c.alignment = center()
            c.border = full_border()

            c = ws.cell(row=row_i, column=best_col)
            if ranks:
                c.value = min(ranks)
                c.font = Font(name="Arial", size=10, bold=True, color=GREEN_UP)
            else:
                c.value = "–"
                c.font = bfont(color=SLATE)
            c.fill = fill(row_bg)
            c.alignment = center()
            c.border = full_border()

        ws.freeze_panes = "B5"

        csr = len(keywords) + 8
        ws.merge_cells(start_row=csr, start_column=1, end_row=csr, end_column=len(dates) + 1)
        c = ws.cell(row=csr, column=1)
        c.value = "Chart Data – Average Rank per Week"
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
        c.fill = fill(ACCENT_BLUE)

        for col_i, label in [(1, "Week"), (2, "Avg Rank"), (3, "# In Top 10")]:
            c = ws.cell(row=csr + 1, column=col_i)
            c.value = label
            c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
            c.fill = fill(DARK_BG)
            c.alignment = center()

        for d_i, d in enumerate(dates):
            r = csr + 2 + d_i
            d_data = loc_df[loc_df['date'] == d]
            avg_r  = d_data['top_rank__avg__total__value'].mean() if len(d_data) else None
            t10    = (d_data['top_rank__avg__total__value'] <= 10).sum() if len(d_data) else 0
            ws.cell(row=r, column=1).value = d.strftime("%b %d, %Y")
            ws.cell(row=r, column=1).alignment = center()
            ws.cell(row=r, column=2).value = round(float(avg_r), 1) if avg_r else None
            ws.cell(row=r, column=2).alignment = center()
            ws.cell(row=r, column=3).value = int(t10)
            ws.cell(row=r, column=3).alignment = center()

        cats     = Reference(ws, min_col=1, min_row=csr + 2, max_row=csr + 1 + len(dates))
        data_ref = Reference(ws, min_col=2, max_col=2, min_row=csr + 1, max_row=csr + 1 + len(dates))
        bar_ref  = Reference(ws, min_col=3, max_col=3, min_row=csr + 1, max_row=csr + 1 + len(dates))

        lc = LineChart()
        lc.title = f"Avg Keyword Rank Over Time – {city}"
        lc.style = 10
        lc.y_axis.title = "Avg Rank (lower = better)"
        lc.x_axis.title = "Week"
        lc.height, lc.width = 12, 24
        lc.add_data(data_ref, titles_from_data=True)
        lc.set_categories(cats)
        lc.series[0].graphicalProperties.line.solidFill = ACCENT_BLUE
        lc.series[0].graphicalProperties.line.width = 25000
        lc.series[0].marker.symbol = "circle"
        lc.series[0].marker.size   = 6
        ws.add_chart(lc, f"A{csr + len(dates) + 4}")

        bc = BarChart()
        bc.title = f"Keywords in Top 10 – {city}"
        bc.style, bc.type = 10, "col"
        bc.y_axis.title = "# Keywords"
        bc.x_axis.title = "Week"
        bc.height, bc.width = 12, 24
        bc.add_data(bar_ref, titles_from_data=True)
        bc.set_categories(cats)
        bc.series[0].graphicalProperties.solidFill = GREEN_UP
        ws.add_chart(bc, f"M{csr + len(dates) + 4}")

    # ── RAW DATA ───────────────────────────────────────────────────────────
    raw_ws = wb.create_sheet("📋 Raw Data")
    raw_ws.sheet_properties.tabColor = SLATE

    keep_cols = ['requested','phrase','location','device','engine',
                 'top_rank__avg__total__value','top_rank__avg__total__change',
                 'top_rank__avg__total__best','top_rank__avg__total__worst',
                 'result__url__url','result__url__domain']
    available = [c for c in keep_cols if c in df.columns]
    rename_map = {
        'requested':'Date','phrase':'Keyword','location':'Location',
        'device':'Device','engine':'Engine',
        'top_rank__avg__total__value':'Rank','top_rank__avg__total__change':'Change',
        'top_rank__avg__total__best':'Best Rank','top_rank__avg__total__worst':'Worst Rank',
        'result__url__url':'Ranking URL','result__url__domain':'Domain',
    }
    raw_df = df[available].copy()
    raw_df.rename(columns={k: v for k, v in rename_map.items() if k in raw_df.columns}, inplace=True)
    raw_df['Date'] = pd.to_datetime(raw_df['Date']).dt.strftime("%Y-%m-%d")
    raw_df = raw_df.sort_values(['Location','Keyword','Date'])

    raw_ws.row_dimensions[1].height = 28
    col_ws_raw = [14, 38, 30, 10, 10, 8, 10, 10, 12, 55, 28]
    for col_i, col_name in enumerate(raw_df.columns, start=1):
        c = raw_ws.cell(row=1, column=col_i)
        c.value = col_name
        c.font = hfont()
        c.fill = fill(DARK_BG)
        c.alignment = center()
        raw_ws.column_dimensions[get_column_letter(col_i)].width = col_ws_raw[col_i-1] if col_i <= len(col_ws_raw) else 14

    for row_i, (_, row) in enumerate(raw_df.iterrows(), start=2):
        row_bg = WHITE if row_i % 2 == 0 else GRAY_BG
        raw_ws.row_dimensions[row_i].height = 16
        for col_i, val in enumerate(row, start=1):
            c = raw_ws.cell(row=row_i, column=col_i)
            col_name = raw_df.columns[col_i - 1]
            if col_name == 'Ranking URL' and pd.notna(val) and str(val).startswith('http'):
                c.value = str(val)
                c.hyperlink = str(val)
                c.font = Font(name="Arial", size=9, color="1D4ED8", underline="single")
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.value = val
                c.font = bfont(size=9)
                c.alignment = center() if col_i not in (2, 3, 10) else left_align()
            c.fill = fill(row_bg)
            c.border = full_border()

    raw_ws.freeze_panes = "A2"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_domain = client_domain.replace('.','_').replace('/','_') if client_domain else "report"
    filename = f"RYNO_{safe_domain}_{latest_date}_keyword_report.xlsx"
    return output, filename


# ── HTML Template ─────────────────────────────────────────────────────────
HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>RYNO – Keyword Report Generator</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500&display=swap" rel="stylesheet"/>
<style>
  :root{--bg:#0d1117;--surface:#161b22;--surface2:#1e2631;--border:#30363d;--blue:#2563eb;--blue-glow:#3b82f6;--teal:#0ea5e9;--green:#16a34a;--red:#dc2626;--amber:#d97706;--text:#e6edf3;--muted:#7d8590;--accent:#58a6ff;--radius:12px}
  *{box-sizing:border-box;margin:0;padding:0}
  body{background:var(--bg);color:var(--text);font-family:'DM Sans',sans-serif;min-height:100vh;overflow-x:hidden}
  body::before{content:'';position:fixed;inset:0;background-image:linear-gradient(rgba(37,99,235,.04) 1px,transparent 1px),linear-gradient(90deg,rgba(37,99,235,.04) 1px,transparent 1px);background-size:40px 40px;pointer-events:none;z-index:0}
  .orb{position:fixed;border-radius:50%;filter:blur(120px);pointer-events:none;z-index:0}
  .orb-1{width:600px;height:600px;background:rgba(37,99,235,.12);top:-200px;right:-100px}
  .orb-2{width:400px;height:400px;background:rgba(14,165,233,.08);bottom:-100px;left:-100px}
  .wrap{position:relative;z-index:1;max-width:860px;margin:0 auto;padding:60px 24px 80px}
  .header{text-align:center;margin-bottom:52px}
  .badge{display:inline-flex;align-items:center;gap:8px;background:rgba(37,99,235,.15);border:1px solid rgba(37,99,235,.35);border-radius:100px;padding:6px 16px;font-family:'DM Mono',monospace;font-size:12px;color:var(--accent);letter-spacing:.08em;text-transform:uppercase;margin-bottom:24px}
  .badge span{width:7px;height:7px;background:var(--blue-glow);border-radius:50%;animation:pulse 2s infinite}
  @keyframes pulse{0%,100%{opacity:1;transform:scale(1)}50%{opacity:.5;transform:scale(.8)}}
  h1{font-family:'Syne',sans-serif;font-size:clamp(30px,5vw,52px);font-weight:800;line-height:1.1;background:linear-gradient(135deg,#fff 30%,#7dd3fc 100%);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;margin-bottom:16px}
  .sub{color:var(--muted);font-size:16px;font-weight:300;max-width:480px;margin:0 auto;line-height:1.6}
  .upload-area{border:2px dashed var(--border);border-radius:20px;background:var(--surface);padding:64px 40px;text-align:center;cursor:pointer;transition:all .25s;position:relative;overflow:hidden}
  .upload-area::before{content:'';position:absolute;inset:0;background:radial-gradient(ellipse at center,rgba(37,99,235,.06),transparent 70%);opacity:0;transition:opacity .3s}
  .upload-area:hover,.upload-area.drag{border-color:var(--blue);background:rgba(37,99,235,.06);transform:translateY(-2px);box-shadow:0 20px 60px rgba(37,99,235,.15),0 0 0 1px rgba(37,99,235,.2)}
  .upload-area:hover::before,.upload-area.drag::before{opacity:1}
  .upload-icon{width:72px;height:72px;background:linear-gradient(135deg,rgba(37,99,235,.2),rgba(14,165,233,.2));border:1px solid rgba(37,99,235,.3);border-radius:20px;display:flex;align-items:center;justify-content:center;margin:0 auto 24px;font-size:32px;transition:transform .2s}
  .upload-area:hover .upload-icon{transform:scale(1.1) rotate(-3deg)}
  .upload-title{font-family:'Syne',sans-serif;font-size:20px;font-weight:700;margin-bottom:8px}
  .upload-sub{color:var(--muted);font-size:14px;margin-bottom:24px}
  .btn{display:inline-flex;align-items:center;gap:8px;background:var(--blue);color:#fff;border:none;border-radius:10px;padding:12px 28px;font-family:'DM Sans',sans-serif;font-size:15px;font-weight:500;cursor:pointer;transition:all .2s;box-shadow:0 4px 20px rgba(37,99,235,.4)}
  .btn:hover{background:#1d4ed8;transform:translateY(-1px);box-shadow:0 6px 28px rgba(37,99,235,.5)}
  #csv-input{display:none}
  .file-info{display:none;background:var(--surface2);border:1px solid var(--border);border-radius:var(--radius);padding:20px 24px;margin-top:20px;align-items:center;gap:16px}
  .file-info.show{display:flex}
  .file-icon{width:48px;height:48px;background:rgba(22,163,74,.15);border:1px solid rgba(22,163,74,.3);border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:22px;flex-shrink:0}
  .file-det{flex:1;min-width:0}
  .file-name{font-weight:500;font-size:14px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
  .file-meta{color:var(--muted);font-size:12px;font-family:'DM Mono',monospace;margin-top:3px}
  .client-tag{display:none;align-items:center;gap:12px;background:rgba(37,99,235,.1);border:1px solid rgba(37,99,235,.25);border-radius:var(--radius);padding:14px 20px;margin-top:14px}
  .client-tag.show{display:flex}
  .client-lbl{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;font-family:'DM Mono',monospace}
  .client-dom{font-family:'DM Mono',monospace;font-size:15px;color:var(--accent);font-weight:500;margin-top:2px}
  .stats{display:none;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin-top:14px}
  .stats.show{display:grid}
  .stat{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:18px 20px;text-align:center;position:relative;overflow:hidden}
  .stat::after{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,var(--blue),var(--teal));border-radius:2px 2px 0 0}
  .stat-v{font-family:'Syne',sans-serif;font-size:26px;font-weight:800;color:#fff;line-height:1;margin-bottom:6px}
  .stat-l{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.06em}
  .gen-wrap{display:none;margin-top:28px;text-align:center}
  .gen-wrap.show{display:block}
  .btn-gen{display:inline-flex;align-items:center;gap:12px;background:linear-gradient(135deg,var(--blue),var(--teal));color:#fff;border:none;border-radius:14px;padding:18px 52px;font-family:'Syne',sans-serif;font-size:18px;font-weight:700;cursor:pointer;transition:all .25s;box-shadow:0 8px 32px rgba(37,99,235,.4);position:relative;overflow:hidden}
  .btn-gen::before{content:'';position:absolute;inset:0;background:linear-gradient(135deg,rgba(255,255,255,.15),transparent 50%)}
  .btn-gen:hover:not(:disabled){transform:translateY(-3px);box-shadow:0 16px 48px rgba(37,99,235,.5)}
  .btn-gen:disabled{opacity:.5;cursor:not-allowed;transform:none}
  .progress{display:none;margin-top:24px;background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:28px}
  .progress.show{display:block}
  .prog-hd{display:flex;justify-content:space-between;margin-bottom:10px;font-size:13px;color:var(--muted)}
  .prog-track{background:var(--surface2);border-radius:100px;height:8px;overflow:hidden}
  .prog-fill{height:100%;background:linear-gradient(90deg,var(--blue),var(--teal));border-radius:100px;width:0%;transition:width .5s ease}
  .steps{margin-top:20px;display:flex;flex-direction:column;gap:10px}
  .step{display:flex;align-items:center;gap:10px;font-size:13px;color:var(--muted);transition:color .2s}
  .step.done{color:var(--green)}.step.active{color:var(--text)}
  .dot{width:18px;height:18px;border-radius:50%;border:2px solid currentColor;display:flex;align-items:center;justify-content:center;font-size:10px;flex-shrink:0}
  .step.done .dot{background:var(--green);border-color:var(--green)}
  .step.active .dot{border-color:var(--blue);animation:spin 1s linear infinite}
  @keyframes spin{to{transform:rotate(360deg)}}
  .success{display:none;background:linear-gradient(135deg,rgba(22,163,74,.1),rgba(14,165,233,.1));border:1px solid rgba(22,163,74,.3);border-radius:20px;padding:44px;text-align:center;margin-top:24px;animation:up .4s ease}
  .success.show{display:block}
  @keyframes up{from{opacity:0;transform:translateY(16px)}to{opacity:1;transform:translateY(0)}}
  .s-icon{font-size:56px;margin-bottom:16px}
  .s-title{font-family:'Syne',sans-serif;font-size:24px;font-weight:800;margin-bottom:8px;color:#fff}
  .s-sub{color:var(--muted);font-size:14px;margin-bottom:6px}
  .s-file{font-family:'DM Mono',monospace;font-size:12px;color:var(--accent);margin-bottom:28px}
  .btn-dl{display:inline-flex;align-items:center;gap:10px;background:var(--green);color:#fff;border:none;border-radius:12px;padding:16px 36px;font-family:'Syne',sans-serif;font-size:16px;font-weight:700;cursor:pointer;transition:all .2s;box-shadow:0 6px 24px rgba(22,163,74,.4);text-decoration:none}
  .btn-dl:hover{background:#15803d;transform:translateY(-2px);box-shadow:0 10px 32px rgba(22,163,74,.5)}
  .btn-again{display:inline-flex;align-items:center;gap:8px;background:none;border:1px solid var(--border);color:var(--muted);border-radius:12px;padding:14px 28px;font-size:14px;cursor:pointer;margin-left:12px;transition:all .2s;font-family:'DM Sans',sans-serif}
  .btn-again:hover{border-color:var(--accent);color:var(--accent)}
  .error{display:none;background:rgba(220,38,38,.1);border:1px solid rgba(220,38,38,.3);border-radius:var(--radius);padding:16px 20px;color:#fca5a5;font-size:14px;margin-top:16px}
  .error.show{display:block}
  .cards{margin-top:52px;display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:16px}
  .card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:24px;transition:border-color .2s}
  .card:hover{border-color:rgba(37,99,235,.4)}
  .card-n{font-family:'DM Mono',monospace;font-size:11px;color:var(--blue-glow);letter-spacing:.1em;text-transform:uppercase;margin-bottom:10px}
  .card-t{font-family:'Syne',sans-serif;font-weight:700;font-size:15px;margin-bottom:8px;color:#fff}
  .card-b{font-size:13px;color:var(--muted);line-height:1.6}
  .quality-badge{display:inline-flex;align-items:center;gap:6px;background:rgba(22,163,74,.15);border:1px solid rgba(22,163,74,.3);border-radius:8px;padding:4px 12px;font-size:12px;color:#86efac;font-family:'DM Mono',monospace;margin-top:12px}
</style>
</head>
<body>
<div class="orb orb-1"></div>
<div class="orb orb-2"></div>
<div class="wrap">
  <div class="header">
    <div class="badge"><span></span> RYNO SEO Tools</div>
    <h1>Keyword Report<br>Generator</h1>
    <p class="sub">Drop any client's ranking CSV and get a fully styled Excel report — color-coded ranks, charts, trend analysis, and ranking URLs.</p>
    <div class="quality-badge">✦ Identical to the Python-generated report</div>
  </div>

  <form id="upload-form" method="POST" action="/generate" enctype="multipart/form-data">
    <div class="upload-area" id="drop-area" onclick="document.getElementById('csv-input').click()">
      <div class="upload-icon">📂</div>
      <div class="upload-title">Drop your CSV here</div>
      <div class="upload-sub">Drag & drop or click to browse — works for any client account</div>
      <button type="button" class="btn" onclick="event.stopPropagation();document.getElementById('csv-input').click()">Browse Files</button>
      <input type="file" id="csv-input" name="csv_file" accept=".csv"/>
    </div>
  </form>

  <div class="file-info" id="file-info">
    <div class="file-icon">📄</div>
    <div class="file-det">
      <div class="file-name" id="file-name"></div>
      <div class="file-meta" id="file-meta"></div>
    </div>
  </div>

  <div class="client-tag" id="client-tag">
    <div style="font-size:22px">🌐</div>
    <div>
      <div class="client-lbl">Client Domain Detected</div>
      <div class="client-dom" id="client-dom">—</div>
    </div>
  </div>

  <div class="stats" id="stats">
    <div class="stat"><div class="stat-v" id="s-kw">—</div><div class="stat-l">Keywords</div></div>
    <div class="stat"><div class="stat-v" id="s-loc">—</div><div class="stat-l">Locations</div></div>
    <div class="stat"><div class="stat-v" id="s-wk">—</div><div class="stat-l">Weeks</div></div>
    <div class="stat"><div class="stat-v" id="s-rows">—</div><div class="stat-l">Data Rows</div></div>
  </div>

  <div class="error" id="error"></div>

  <div class="gen-wrap" id="gen-wrap">
    <button class="btn-gen" id="btn-gen" onclick="generate()">⚡ Generate Styled Excel Report</button>
  </div>

  <div class="progress" id="progress">
    <div class="prog-hd"><span id="prog-txt">Building report…</span><span id="prog-pct">0%</span></div>
    <div class="prog-track"><div class="prog-fill" id="prog-fill"></div></div>
    <div class="steps">
      <div class="step" id="s1"><div class="dot">→</div><span>Parsing CSV & detecting client domain</span></div>
      <div class="step" id="s2"><div class="dot">→</div><span>Building Dashboard with KPI cards</span></div>
      <div class="step" id="s3"><div class="dot">→</div><span>Creating color-coded location sheets</span></div>
      <div class="step" id="s4"><div class="dot">→</div><span>Adding charts & ranking URLs</span></div>
      <div class="step" id="s5"><div class="dot">→</div><span>Finalizing & packaging Excel file</span></div>
    </div>
  </div>

  <div class="success" id="success">
    <div class="s-icon">🎉</div>
    <div class="s-title">Report Ready!</div>
    <div class="s-sub" id="s-sub"></div>
    <div class="s-file" id="s-file"></div>
    <a class="btn-dl" id="btn-dl" href="#">⬇ Download Excel Report</a>
    <button class="btn-again" onclick="reset()">↩ Generate Another</button>
  </div>

  <div class="cards">
    <div class="card">
      <div class="card-n">Step 01</div>
      <div class="card-t">Export your CSV</div>
      <div class="card-b">Download the keyword ranking CSV from your SEO platform. Works for any client — domain is auto-detected.</div>
    </div>
    <div class="card">
      <div class="card-n">Step 02</div>
      <div class="card-t">Drop & Generate</div>
      <div class="card-b">Drag the CSV here and hit Generate. Gets the exact same styled report as the Python script — colors, charts, everything.</div>
    </div>
    <div class="card">
      <div class="card-n">Step 03</div>
      <div class="card-t">Download & Share</div>
      <div class="card-b">Download the fully formatted Excel file named after the client's domain. Ready to send immediately.</div>
    </div>
  </div>
</div>

<script>
let selectedFile = null, downloadUrl = null, outputFilename = null;

const drop = document.getElementById('drop-area');
drop.addEventListener('dragover', e => { e.preventDefault(); drop.classList.add('drag'); });
drop.addEventListener('dragleave', () => drop.classList.remove('drag'));
drop.addEventListener('drop', e => { e.preventDefault(); drop.classList.remove('drag'); const f = e.dataTransfer.files[0]; if (f) setFile(f); });
document.getElementById('csv-input').addEventListener('change', e => { if (e.target.files[0]) setFile(e.target.files[0]); });

function setFile(file) {
  if (!file.name.endsWith('.csv')) { showErr('Please upload a CSV file.'); return; }
  hideErr();
  selectedFile = file;

  document.getElementById('file-name').textContent = file.name;
  document.getElementById('file-meta').textContent = `${(file.size/1024).toFixed(1)} KB`;
  document.getElementById('file-info').classList.add('show');

  // Quick parse preview
  const reader = new FileReader();
  reader.onload = e => previewCSV(e.target.result, file.size);
  reader.readAsText(file);
}

function parseCSVLine(line) {
  const r=[]; let c='', q=false;
  for (let i=0;i<line.length;i++) {
    if(line[i]==='"'){q=!q;}
    else if(line[i]===','&&!q){r.push(c);c='';}
    else{c+=line[i];}
  }
  r.push(c); return r;
}

function previewCSV(text, size) {
  const lines = text.trim().split('\\n');
  const headers = parseCSVLine(lines[0]).map(h=>h.trim());
  const rows = [];
  for (let i=1;i<Math.min(lines.length,100);i++) {
    if(!lines[i].trim()) continue;
    const vals=parseCSVLine(lines[i]);
    const row={};
    headers.forEach((h,idx)=>{row[h]=(vals[idx]||'').trim();});
    rows.push(row);
  }

  // Parse all for stats
  const allRows = [];
  for (let i=1;i<lines.length;i++) {
    if(!lines[i].trim()) continue;
    const vals=parseCSVLine(lines[i]);
    const row={};
    headers.forEach((h,idx)=>{row[h]=(vals[idx]||'').trim();});
    allRows.push(row);
  }

  const locs = [...new Set(allRows.map(r=>r.location).filter(Boolean))];
  const kws  = [...new Set(allRows.map(r=>r.phrase).filter(Boolean))];
  const dates= [...new Set(allRows.map(r=>(r.requested||'').split(' ')[0]).filter(Boolean))];

  let domain = '';
  if (headers.includes('result__url__domain')) {
    const doms = [...new Set(allRows.map(r=>r['result__url__domain']).filter(Boolean))];
    if (doms.length) domain = doms[0];
  }

  if (domain) {
    document.getElementById('client-dom').textContent = domain;
    document.getElementById('client-tag').classList.add('show');
  }

  document.getElementById('s-kw').textContent   = kws.length * locs.length;
  document.getElementById('s-loc').textContent  = locs.length;
  document.getElementById('s-wk').textContent   = dates.length;
  document.getElementById('s-rows').textContent = allRows.length.toLocaleString();
  document.getElementById('stats').classList.add('show');
  document.getElementById('gen-wrap').classList.add('show');
  document.getElementById('success').classList.remove('show');
  document.getElementById('progress').classList.remove('show');
}

async function generate() {
  if (!selectedFile) return;
  document.getElementById('btn-gen').disabled = true;
  document.getElementById('gen-wrap').classList.remove('show');
  document.getElementById('progress').classList.add('show');
  hideErr();

  const steps = ['s1','s2','s3','s4','s5'];
  const pcts  = [10, 25, 55, 80, 95];
  const msgs  = ['Parsing CSV…','Building Dashboard…','Creating location sheets…','Adding charts & URLs…','Packaging Excel…'];
  let si = 0;

  function advStep(i) {
    if (i > 0) setStep(steps[i-1], 'done');
    if (i < steps.length) setStep(steps[i], 'active');
    setProgress(pcts[i], msgs[i]);
  }

  advStep(0);
  const formData = new FormData();
  formData.append('csv_file', selectedFile);

  // Simulate step progression while waiting
  const timer = setInterval(() => {
    si++;
    if (si < steps.length - 1) advStep(si);
  }, 800);

  try {
    const resp = await fetch('/generate', { method:'POST', body: formData });
    clearInterval(timer);

    if (!resp.ok) {
      const err = await resp.text();
      throw new Error(err || 'Server error');
    }

    // Get filename from header
    const cd = resp.headers.get('Content-Disposition') || '';
    const fnMatch = cd.match(/filename="([^"]+)"/);
    outputFilename = fnMatch ? fnMatch[1] : 'keyword_report.xlsx';

    const blob = await resp.blob();
    downloadUrl = URL.createObjectURL(blob);

    steps.forEach(s => setStep(s, 'done'));
    setProgress(100, 'Complete!');
    await new Promise(r => setTimeout(r, 400));

    document.getElementById('progress').classList.remove('show');
    document.getElementById('success').classList.add('show');
    document.getElementById('s-sub').textContent = `Fully styled with color-coded ranks, charts & ranking URLs`;
    document.getElementById('s-file').textContent = '📁 ' + outputFilename;
    document.getElementById('btn-dl').href = downloadUrl;
    document.getElementById('btn-dl').download = outputFilename;
  } catch(err) {
    clearInterval(timer);
    showErr('Error generating report: ' + err.message);
    document.getElementById('btn-gen').disabled = false;
    document.getElementById('gen-wrap').classList.add('show');
    document.getElementById('progress').classList.remove('show');
  }
}

function setProgress(pct, txt) {
  document.getElementById('prog-fill').style.width = pct+'%';
  document.getElementById('prog-pct').textContent = pct+'%';
  if (txt) document.getElementById('prog-txt').textContent = txt;
}
function setStep(id, state) {
  const el = document.getElementById(id);
  el.classList.remove('done','active');
  if (state) el.classList.add(state);
  el.querySelector('.dot').textContent = state==='done' ? '✓' : '→';
}
function showErr(msg) { const e=document.getElementById('error'); e.textContent=msg; e.classList.add('show'); }
function hideErr() { document.getElementById('error').classList.remove('show'); }
function reset() {
  selectedFile=null; downloadUrl=null; outputFilename=null;
  ['file-info','client-tag','stats','gen-wrap','progress','success'].forEach(id=>document.getElementById(id).classList.remove('show'));
  document.getElementById('csv-input').value='';
  document.getElementById('btn-gen').disabled=false;
  hideErr();
}
</script>
</body>
</html>"""


@app.route('/')
def index():
    return render_template_string(HTML)


@app.route('/generate', methods=['POST'])
def generate():
    from flask import Response
    if 'csv_file' not in request.files:
        return Response('No file uploaded', status=400)
    f = request.files['csv_file']
    if not f.filename.endswith('.csv'):
        return Response('Please upload a CSV file', status=400)
    try:
        df = pd.read_csv(f, low_memory=False)
        required = ['requested', 'phrase', 'location', 'top_rank__avg__total__value']
        missing = [c for c in required if c not in df.columns]
        if missing:
            return Response(f'Missing required columns: {", ".join(missing)}', status=400)
        output, filename = generate_excel(df, original_filename=f.filename)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return Response(f'Error: {str(e)}', status=500)


if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5050))
    print("\n" + "="*55)
    print("  🔍  RYNO SEO Tools – Keyword Report Generator")
    print("="*55)
    print(f"  Open in browser: http://localhost:{port}")
    print("  Press Ctrl+C to stop")
    print("="*55 + "\n")
    app.run(debug=False, host="0.0.0.0", port=port)
